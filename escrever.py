#não sei
import re

#excel
from tkinter.font import Font
from numpy import average
import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

#beautifulsoup
from bs4 import BeautifulSoup
import requests
from sqlalchemy import null
from sympy import block_collapse

#statistics
import statistics

wb = Workbook()
ws = wb.active
ws.title = "escrever"

#--------------------DADOS--------------------

#dados inseridos com as profissões de cada curso e tudo mais
dados = {
    "Ciências Biológicas":{
        "Biomedicina":"Biomédico",
        "Biologia":"Biólogo",
        "Farmácia":"Farmacêutico",
        "Ecologia":"Ecólogo",
        "Enfermagem":"Enfermeiro",
        "Engenharia Agronômica":"Engenheiro Agrônomo",
        "Engenharia de Pesca": "Engenheiro de Pesca",
        "Engenharia Florestal": "Engenheiro Florestal",
        "Fisioterapia":"Fisioterapeuta",
        "Fonoaudiologia":"Fonoaudiólogo",
        "Medicina":"Médico",
        "Medicina Veterinária": "Médico Veterinário",
        "Nutrição": "Nutricionista",
        "Odontologia": "Dentista",
        "Terapia Ocupacional": "Terapeuta Ocupacional",
        "Zootecnia": "Zootecnista", 
    },
    "Ciências Exatas":{
        "Engenharia Aeronáutica":"Engenheiro Aeronáutico",
        "Engenharia Ambiental": "Engenheiro Ambiental",
        "Engenharia Cartográfica e de Agrimensura": "Engenheiro Cartografo",
        "Engenharia Civil":"Engenheiro Civil",
        "Engenharia de Alimentos":"Engenheiro de Alimentos",
        "Engenharia de Bioprocessos e Biotecnologia":"Biotecnólogo",
        "Engenharia de Controle e Automação": "Engenheiro de Automação",
        "Engenharia de Eletrônica Telecomunicações": "Engenheiro de Telecomunicações",
        "Engenharia de Materiais ": "Engenheiro de Materiais",
        "Engenharia de Produção ": "Engenheiro de Produção",
        "Engenharia Elétrica ": "Engenheiro Elétrico",
        "Engenharia Industrial Madeireiro ": "Engenheiro Especialista em Tecnologia de Madeira",
        "Engenharia Mecânica ": "Engenheiro Mecânico",
        "Engenharia Quimica ": "Engenheiro Químico",
        "Estatistica ": "Estatístico",
        "Física":"Físico",
        "Física Médica": "Físico Médico",
        "Geologia":"Geólogo",
        "Matemática": "Matemático",
        "Metereologia": "Meteorologista",
        "Química": "Químico",
    },
    "Ciências Humanas":{
        "Administrador":"Administrador",
        "Administração Pública":"Administrador Público",
        "Arquitetura e Urbanismo":"Arquiteto",
        "Arquivologia":"Arquivista",
        "Artes Visuais":"Artista",
        "Arte-Teatro e Artes Cênicas":"Ator",
        "Biblioteconomia":"Biblioteconomista",
        "Ciências Econômicas":"Economista",
        "Ciências Sociais":"Sociólogo",
        "Comunicação: Rádio, TV e Internet": "Telecomunicações",
        "Design":"Designer",
        "Direito":"Advogado",
        "Filosofia":"Professor de Filosofia",
        "Geografia":"Geografo",
        "História":"Historiador",
        "Jornalismo":"Jornalista",
        "Letras":"Linguista",
        "Letras-Tradutor":"Tradutor",
        "Música":"Músico",
        "Pedagogo":"Pedagogo",
        "Psicologia":"Psicólogo",
        "Relações Internacionais":"Agente de Comércio Exterior",
        "Relações Públicas":"Relações Publicas",
        "Serviço Social":"Assistente Social",
        "Turismo":"Guia de Turismo",
        
    }
    
    
}
newdados = {}


#-----------------WEBSCRAPING-----------------------
#pesquisa dentro dos dados e faz loop para cada profissão
for ciencia in dados:
    for curso in dados[ciencia]:
        #puxar a sopa
        profissao_dados = str(dados[ciencia][curso])
        primeira_letra = profissao_dados[0]
        #pesquisa o site com a primeira letra da profissão na url
        html_text = requests.get('https://www.salario.com.br/tabela-salarial/?cargos='+primeira_letra+'#listaSalarial').text
        soup = BeautifulSoup(html_text, 'lxml')
        #procurar na tabela de profissões
        tabela_sopa = soup.find('table', class_="listas")
        #procurar a profissão especifica
        profissoes = tabela_sopa.find_all('td',attrs={'data-label': "Cargo"})
        
        #define as listas
        lista_piso = []
        lista_teto = []
        for profissao in profissoes:
            profissao_cargo = profissao.text
            profissao_pai = profissao.parent
            piso_salarial = profissao_pai.find('td', attrs={'data-label':"Piso Salarial"})
            teto_salarial = profissao_pai.find('td', attrs={'data-label':"Teto Salarial"})
            if(profissao_dados in profissao_cargo):
                #print(piso_salarial.text)
                #print(teto_salarial.text)
                
                #pega o valor dos salários e adiciona na lista
                lista_teto.append(float(teto_salarial.text.replace(".",'').replace(",",".")))
                lista_piso.append(float(piso_salarial.text.replace(".",'').replace(",",".")))
                
        #apenas calcular as médias que tenham pelo menos algum valor para calcular
        if((len(lista_teto)>0) and len(lista_piso)>0):
            
            #calcular a média dos salários
            media_teto = round(statistics.mean(lista_teto),2)
            media_piso = round(statistics.mean(lista_piso),2)   
        #se não houver. zerar as médias
        else:
            media_teto = 0
            media_piso = 0
        
        #cria um novo dicionário com os valores prontos para o EXCEL
        newdados.update({profissao_dados:{"piso":media_piso,"teto":media_teto,"curso":curso}})
            
        #reseta as listas
        lista_piso = []
        lista_teto = []
       
        
        

        
#------------------EXCEL-------------------
#criar o cabeçalho do excel
ws.append(["Curso","Profissão","Piso salarial","Teto salarial"])
#loop de cada uma das profissões dos dados
for key in newdados:
    profissao_excel = str(key)
    piso_excel = newdados[key]['piso']
    teto_excel = newdados[key]['teto']
    curso_excel = newdados[key]['curso']
    #colocar os dados na tabela do excel
    ws.append([curso_excel,profissao_excel,piso_excel,teto_excel])
        #print(profissao +" "+ licenciatura)


wb.save('escrever.xlsx')
#excel
from tkinter.font import Font
from numpy import average
import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

wl = load_workbook('escrever.xlsx',data_only=True)
sh = wl['escrever']

for cell in sh['B']:
    if(cell.value=="Ciências Biológicas"):
        cell.fill = PatternFill("solid", start_color="bdffd8")
    elif(cell.value=="Ciências Exatas"):
        cell.fill = PatternFill("solid", start_color="ffbdbd")
    elif(cell.value=="Ciências Humanas"):
        cell.fill = PatternFill("solid", start_color="bde8ff")

wl.save('escrever.xlsx')
print("cabou")